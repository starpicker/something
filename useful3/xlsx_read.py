#!/data/data/com.termux/files/usr/bin/python
# coding=utf-8
import openpyxl
import sys

filename = sys.argv[1]
##filename = r'C:\Users\dell\Desktop\1.xlsx'
wb = openpyxl.load_workbook(filename)
sheetName = wb.sheetnames[0]
sheet = wb[sheetName]

col = ord('I') - ord('A') + 1
num = 0
k = 0
for i in range(6, sheet.max_row+1):
    v = sheet.cell(row=i, column=col).value
    if(v is not None and v >= 1):
        print(sheet.cell(row=i, column=1).value, v)
        num += v
        k += 1
print(" 4 档 total %d of %d 种" %  (num, k))
