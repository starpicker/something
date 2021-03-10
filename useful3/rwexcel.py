# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy
from re import findall

wb = load_workbook('1.xlsx')
book = Workbook()

wbsheet=wb.active
wbsheet_new = book.create_sheet()
wm = list(wbsheet.merged_cells)
mergeArr = []

if len(wm)>0 :
    for i in range(0, len(wm)):
        cell2 = str(wm[i]).replace('(<CellRange ','').replace('>,)','')
        mergeArr.append(cell2)
        #wbsheet_new.merge_cells(cell2)
#print(mergeArr)
        
counts = 15
row_start = 1
col_start = 1
row_ends = 29
col_ends = 12
row_span = row_ends-row_start+1
modi_value = 202000220
string = '合肥惠丰电子科技有限公司   IT末梢终端耗材、零配件出库单  NO:YS03 '

for n in range(counts):
    
    for rows in range(row_start, row_ends+1):
        wbsheet_new.row_dimensions[rows+n*row_span].height = wbsheet.row_dimensions[rows].height
        
        for cols in range(col_start, col_ends):
            wbsheet_new.column_dimensions[get_column_letter(cols)].width = wbsheet.column_dimensions[get_column_letter(cols)].width
            target_cell = wbsheet_new.cell(row=rows+n*row_span, column=cols, value = wbsheet.cell(rows, cols).value)
            if (rows == row_start and cols == col_start):
                target_cell = wbsheet_new.cell(row=rows+n*row_span, column=cols, value = string + str(modi_value))
                modi_value += 1
            if (rows == row_start+10 and cols == col_start):
                target_cell = wbsheet_new.cell(row=rows+n*row_span, column=cols, value = string + str(modi_value))
                modi_value += 1
            if (rows == row_start+20 and cols == col_start):
                target_cell = wbsheet_new.cell(row=rows+n*row_span, column=cols, value = string + str(modi_value))
                modi_value += 1
                
            source_cell = wbsheet.cell(rows, cols)

            if wbsheet.cell(rows, cols).has_style:
                target_cell._style          = copy(source_cell._style )
                target_cell.font            = copy(source_cell.font)
                target_cell.border          = copy(source_cell.border)
                target_cell.fill            = copy(source_cell.fill)
                target_cell.number_format   = copy(source_cell.number_format)
                target_cell.protection      = copy(source_cell.protection)
                target_cell.alignment       = copy(source_cell.alignment)
            
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell._hyperlink)
         
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
    
    for i in mergeArr:
        temp1 = findall(r'[A-Z]', i)
        temp2 = findall(r'\d+', i)
        #print(temp1)
        #print(temp2)
        temp3 = [str(int(temp2[0])+n*row_span), str(int(temp2[1])+n*row_span)]
        temp4 = temp1[0]+temp3[0] + ":" + temp1[1]+temp3[1]
        wbsheet_new.merge_cells(temp4)
        
wb.close()
book.save('ml.xlsx')
book.close()
