#!/data/data/com.termux/files/usr/bin/python
# coding=utf-8
import openpyxl
import sys

filename = sys.argv[1]
##filename = r'C:\Users\dell\Desktop\1.xlsx'
wb = openpyxl.load_workbook(filename)
sheetName = wb.sheetnames[0]
sheet = wb[sheetName]
'''
col = ord('I') - ord('A') + 1
num = 0
k = 0
c = 0
tplt = "{0:{3}<15}\t{1:{3}<3}{2:<10}"
times = 0
for i in range(5, sheet.max_row+1):
    v = sheet.cell(row=i, column=col).value
    t = sheet.cell(row=i, column=2).value
    if(v is not None and v >= 1 and v < 100):
        times += 1
        print(tplt.format(sheet.cell(row=i, column=1).value, v, t, chr(12288)))
        num += v
        k += 1
        if "常规" in t:
            c += v
        if times%5 == 0:
            print("\n")
print(" 4 档 total %d(常规:%d, 异型:%d) of %d 种" %  (num, c, num - c, k))
'''

level = 4
col = ord('I') - ord('A') + 1 + level - 4
Lnormal = []
N1 = 0
Labnormal = []
N2 = 0

for i in range(5, sheet.max_row+1):
    v = sheet.cell(row=i, column=col).value
    t = sheet.cell(row=i, column=2).value
    if(v is not None and v >= 1 and v < 100):
        if "常规" in t:
            Lnormal.append([sheet.cell(row=i, column=1).value, v, t])
            N1 += v
        else:
            Labnormal.append([sheet.cell(row=i, column=1).value, v, t])
            N2 += v

def takeNumber(e):
    return e[1]

Lnormal.sort(key=takeNumber, reverse=True)
Labnormal.sort(key=takeNumber, reverse=True)

tplt = "{0:{3}<15}\t{1:{3}<3}{2:<10}"
def printinfo(l):
    times = 0
    for i in l:
        times += 1
        print(tplt.format(i[0], i[1], i[2], chr(12288)))
        if times%5 == 0:
            print("\n")

print("常规(%d):" % (N1))
printinfo(Lnormal)

print("\n")

print("异型(%d):" % (N2))
printinfo(Labnormal)

print("\n")

print("%d 档 total %d条" % (level, N1+N2))
